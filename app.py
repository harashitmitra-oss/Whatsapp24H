import re
import os
import html
from io import BytesIO
from datetime import datetime
from copy import copy

import numpy as np
import pandas as pd
import plotly.express as px
import streamlit as st
from openpyxl.styles import PatternFill, Font, Alignment


# ============================================================
# APP CONFIG
# ============================================================

st.set_page_config(
    page_title="Tetr Community Analysis",
    page_icon="💬",
    layout="wide"
)

APP_TITLE = "Tetr Community Analysis"

STUDENT_PHONE_FILES = [
    "students_phone.xlsx",
    "Students_phone.xlsx",
    "student_phone.xlsx",
    "Students Phone.xlsx",
    "Students_phone.XLSX",
]


# ============================================================
# KEYWORDS
# ============================================================

TOPIC_KEYWORDS = {
    "Admissions": [
        "admission", "apply", "application", "accepted", "admitted",
        "offer letter", "selection", "interview", "seat", "enrol",
        "enroll", "joining", "class of", "cohort"
    ],
    "Payment / Fees": [
        "payment", "paid", "pay", "fee", "fees", "deposit", "refund",
        "invoice", "scholarship", "loan", "deadline", "installment", "emi"
    ],
    "Travel / Campus": [
        "travel", "flight", "visa", "passport", "campus", "country",
        "destination", "india", "uae", "europe", "ghana", "argentina",
        "new york", "nyc", "global", "countries"
    ],
    "Program / Curriculum": [
        "program", "course", "curriculum", "module", "class", "degree",
        "university", "credit", "assignment", "term", "semester",
        "learning", "learn by doing"
    ],
    "Events / Webinars": [
        "event", "webinar", "masterclass", "session", "call", "zoom",
        "meet", "meeting", "orientation", "workshop", "live", "ama",
        "challenge", "competition", "pitch"
    ],
    "Community / Social": [
        "group", "community", "friends", "meetup", "introduction",
        "intro", "fun", "party", "hangout", "connect", "fam",
        "informal group"
    ],
    "Career / Placements": [
        "career", "placement", "job", "internship", "startup",
        "founder", "business", "networking", "mentor", "mentorship",
        "vc", "finance", "consulting", "product", "tech"
    ],
    "Support / Operations": [
        "help", "support", "issue", "problem", "error", "link", "form",
        "document", "documents", "access", "not working", "register",
        "registration"
    ],
    "Challenge / Submission": [
        "submission", "challenge", "ceo", "pitch", "idea", "proposal",
        "solution", "problem", "netflix", "deadline", "winner", "closed"
    ],
}

QUESTION_WORDS = [
    "what", "when", "where", "why", "how", "who", "which",
    "can", "could", "should", "would", "is", "are", "do", "does",
    "did", "will", "any update", "anyone know", "please tell",
    "gonna", "going to", "what time", "which time"
]

HIGH_INTENT_KEYWORDS = [
    "payment", "pay", "paid", "deposit", "confirm", "confirmed",
    "seat", "deadline", "join", "joining", "enroll", "enrol",
    "admission", "offer", "fee", "fees", "scholarship", "loan",
    "call me", "dm me", "interested", "ready", "how to pay",
    "register", "joining tetr", "accepted", "admitted"
]

NEGATIVE_KEYWORDS = [
    "confused", "doubt", "doubts", "worried", "concern", "problem",
    "issue", "not clear", "not sure", "scam", "fake", "expensive",
    "refund", "bad", "delay", "late", "unhappy", "disappointed",
    "doesn't make sense", "not working", "can't", "cannot",
    "no response", "ignored", "lost", "missed", "forgot"
]

POSITIVE_KEYWORDS = [
    "great", "amazing", "excited", "happy", "thanks", "thank you",
    "helpful", "good", "awesome", "love", "perfect", "clear",
    "looking forward", "super", "nice", "congratulations", "congrats",
    "proud", "bravo", "lessgo", "yay", "hyped", "can't wait"
]

ANNOUNCEMENT_KEYWORDS = [
    "announcement", "important", "note", "please note", "reminder",
    "deadline", "update", "everyone", "all", "kindly", "please fill",
    "form", "session", "today", "tomorrow", "join here", "register here",
    "we are live", "we’re live", "live now", "challenge closed",
    "happening today", "minutes to go"
]


# ============================================================
# BASIC HELPERS
# ============================================================

def find_student_phone_file():
    for path in STUDENT_PHONE_FILES:
        if os.path.exists(path):
            return path
    return ""


def clean_col_name(col):
    return str(col).strip().lower().replace("_", " ").replace("-", " ")


def detect_column(df, candidates):
    col_map = {clean_col_name(c): c for c in df.columns}

    for candidate in candidates:
        c = clean_col_name(candidate)
        if c in col_map:
            return col_map[c]

    for original in df.columns:
        cleaned = clean_col_name(original)
        for candidate in candidates:
            c = clean_col_name(candidate)
            if c in cleaned:
                return original

    return None


def clean_name_key(value):
    text = str(value).strip().lower()
    text = text.replace("\u202a", "").replace("\u202c", "")
    text = text.replace("\u200e", "").replace("\u200f", "")
    text = re.sub(r"\s+", " ", text)
    text = text.strip()
    return text


def safe_str(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


# ============================================================
# PHONE MATCHING
# ============================================================

def normalize_phone(value):
    """
    Normalizes many WhatsApp / Excel number formats:
    +91 98765 43210 -> 919876543210
    09876543210 -> 9876543210
    00 91 98765 43210 -> 919876543210
    +1 (825) 533-1215 -> 18255331215
    """
    if pd.isna(value):
        return ""

    text = str(value).strip()

    if not text or text.lower() in ["nan", "none", "null"]:
        return ""

    text = text.replace("\u202a", "").replace("\u202c", "")
    text = text.replace("\u200e", "").replace("\u200f", "")
    digits = re.sub(r"\D+", "", text)

    if not digits:
        return ""

    while digits.startswith("00") and len(digits) > 10:
        digits = digits[2:]

    while digits.startswith("0") and len(digits) > 10:
        digits = digits[1:]

    return digits


def phone_keys(value):
    """
    Generates many possible keys for robust matching.
    """
    digits = normalize_phone(value)

    if not digits:
        return []

    keys = set()

    keys.add(digits)

    no_zero = digits.lstrip("0")
    if no_zero:
        keys.add(no_zero)

    if len(digits) >= 7:
        keys.add(digits[-7:])

    if len(digits) >= 8:
        keys.add(digits[-8:])

    if len(digits) >= 9:
        keys.add(digits[-9:])

    if len(digits) >= 10:
        last10 = digits[-10:]
        keys.add(last10)
        keys.add("91" + last10)
        keys.add("1" + last10)

    if len(digits) >= 11:
        keys.add(digits[-11:])

    if len(digits) >= 12:
        keys.add(digits[-12:])

    if digits.startswith("91") and len(digits) == 12:
        keys.add(digits[2:])

    if digits.startswith("1") and len(digits) == 11:
        keys.add(digits[1:])

    return list(keys)


def extract_phone_from_sender(sender):
    sender = str(sender).strip()
    digits = re.sub(r"\D+", "", sender)

    if len(digits) >= 7:
        return normalize_phone(sender)

    return ""


# ============================================================
# STUDENT / PERSONA WORKBOOK
# ============================================================

def find_name_col(df, entity_type):
    if entity_type == "Persona":
        candidates = [
            "persona name", "persona", "name", "full name", "display name",
            "student name", "candidate name", "lead name", "admin name"
        ]
    else:
        candidates = [
            "student name", "name", "full name", "student", "candidate name",
            "lead name", "display name"
        ]

    return detect_column(df, candidates)


def find_phone_col(df):
    return detect_column(
        df,
        [
            "phone", "phone number", "mobile", "mobile number",
            "whatsapp", "whatsapp number", "whatsapp no", "contact",
            "number", "contact number", "phone no", "wa number",
            "student phone", "persona phone"
        ]
    )


def detect_batch_label(row, df_columns):
    """
    Detects UG B10 / PG B5 from any likely batch/program columns or row text.
    """
    row_text = " ".join(safe_str(row.get(c, "")) for c in df_columns)
    row_text = row_text.replace("_", " ").replace("-", " ")

    direct = re.search(r"\b(UG|PG)\s*B(?:atch)?\s*[- ]?(\d{1,2})\b", row_text, flags=re.I)
    if direct:
        return f"{direct.group(1).upper()} B{direct.group(2)}"

    direct2 = re.search(r"\b(UG|PG)\s*[- ]?B?(\d{1,2})\b", row_text, flags=re.I)
    if direct2:
        return f"{direct2.group(1).upper()} B{direct2.group(2)}"

    batch_num = re.search(r"\bB(?:atch)?\s*[- ]?(\d{1,2})\b", row_text, flags=re.I)
    program = re.search(r"\b(UG|PG)\b", row_text, flags=re.I)

    if program and batch_num:
        return f"{program.group(1).upper()} B{batch_num.group(1)}"

    if batch_num:
        return f"B{batch_num.group(1)}"

    if program:
        return program.group(1).upper()

    return ""


@st.cache_data(show_spinner=False)
def load_people_mapping():
    """
    Sheet 1 = Students
    Sheet 2 = Personas

    Returns:
    phone_mapping, name_mapping, students_df, personas_df, error_message
    """
    file_path = find_student_phone_file()

    if not file_path:
        return {}, {}, pd.DataFrame(), pd.DataFrame(), "students_phone.xlsx not found in repo."

    try:
        xl = pd.ExcelFile(file_path)
    except Exception as e:
        return {}, {}, pd.DataFrame(), pd.DataFrame(), f"Could not read {file_path}: {e}"

    if len(xl.sheet_names) < 1:
        return {}, {}, pd.DataFrame(), pd.DataFrame(), "Workbook has no sheets."

    students_df = pd.read_excel(file_path, sheet_name=xl.sheet_names[0])
    personas_df = pd.DataFrame()

    if len(xl.sheet_names) >= 2:
        personas_df = pd.read_excel(file_path, sheet_name=xl.sheet_names[1])

    phone_mapping = {}
    name_mapping = {}
    errors = []

    def add_rows(source_df, entity_type):
        if source_df is None or source_df.empty:
            return

        name_col = find_name_col(source_df, entity_type)
        phone_col = find_phone_col(source_df)

        if name_col is None:
            errors.append(
                f"{entity_type} sheet: could not detect name column. Found columns: {list(source_df.columns)}"
            )
            return

        for _, row in source_df.iterrows():
            name = safe_str(row.get(name_col, ""))

            if not name or name.lower() in ["nan", "none", "null"]:
                continue

            batch_label = detect_batch_label(row, source_df.columns)

            record = {
                "MatchedName": name,
                "PersonType": entity_type,
                "BatchLabel": batch_label,
                "SourcePhone": "",
                "NameColumn": str(name_col),
                "PhoneColumn": str(phone_col) if phone_col is not None else "",
            }

            name_key = clean_name_key(name)
            if name_key:
                name_mapping[name_key] = record

            if phone_col is not None:
                phone = row.get(phone_col, "")
                record["SourcePhone"] = safe_str(phone)

                for key in phone_keys(phone):
                    if not key:
                        continue

                    if key in phone_mapping:
                        existing = phone_mapping[key]

                        # Prefer Persona over Student only if entity is Persona and existing is not Persona.
                        # Prefer Student if existing is empty/unmatched.
                        if existing.get("PersonType") == "Persona" and entity_type == "Student":
                            continue

                        if existing.get("PersonType") == "Student" and entity_type == "Persona":
                            phone_mapping[key] = record
                        else:
                            phone_mapping[key] = existing
                    else:
                        phone_mapping[key] = record

    add_rows(students_df, "Student")
    add_rows(personas_df, "Persona")

    error_message = " | ".join(errors)

    return phone_mapping, name_mapping, students_df, personas_df, error_message


def match_person(sender, phone_mapping, name_mapping):
    sender = safe_str(sender)
    sender_phone = extract_phone_from_sender(sender)

    if sender_phone:
        for key in phone_keys(sender_phone):
            if key in phone_mapping:
                rec = phone_mapping[key]
                return {
                    "MatchedName": rec.get("MatchedName", ""),
                    "PersonType": rec.get("PersonType", "Matched"),
                    "BatchLabel": rec.get("BatchLabel", ""),
                    "SenderPhone": sender_phone,
                    "IsMatched": True,
                    "IsPersona": rec.get("PersonType") == "Persona"
                }

    sender_key = clean_name_key(sender)

    if sender_key in name_mapping:
        rec = name_mapping[sender_key]
        return {
            "MatchedName": rec.get("MatchedName", ""),
            "PersonType": rec.get("PersonType", "Matched"),
            "BatchLabel": rec.get("BatchLabel", ""),
            "SenderPhone": sender_phone,
            "IsMatched": True,
            "IsPersona": rec.get("PersonType") == "Persona"
        }

    return {
        "MatchedName": "",
        "PersonType": "Unmatched",
        "BatchLabel": "",
        "SenderPhone": sender_phone,
        "IsMatched": False,
        "IsPersona": False
    }


def enrich_with_people(df, phone_mapping, name_mapping):
    df = df.copy()

    matched = df["Sender"].apply(lambda x: match_person(x, phone_mapping, name_mapping)).apply(pd.Series)

    df["SenderPhone"] = matched["SenderPhone"]
    df["MatchedName"] = matched["MatchedName"]
    df["PersonType"] = matched["PersonType"]
    df["BatchLabel"] = matched["BatchLabel"]
    df["IsMatched"] = matched["IsMatched"]
    df["IsPersona"] = matched["IsPersona"]

    df["DisplayName"] = np.where(
        df["MatchedName"].astype(str).str.strip() != "",
        df["MatchedName"],
        df["Sender"]
    )

    return df


def infer_community_names(df):
    """
    Rename each uploaded file community based on dominant matched UG/PG batch.
    """
    df = df.copy()
    rename_map = {}
    used = {}

    for original_community, g in df.groupby("OriginalCommunity"):
        candidates = g[
            (~g["IsSystem"])
            & (g["BatchLabel"].astype(str).str.strip() != "")
            & (g["IsMatched"])
        ].copy()

        if candidates.empty:
            new_name = original_community
        else:
            unique_people = candidates.drop_duplicates(
                subset=["PersonType", "MatchedName", "SenderPhone", "BatchLabel"]
            )

            counts = unique_people["BatchLabel"].value_counts()

            if counts.empty:
                new_name = original_community
            else:
                new_name = counts.index[0]

        if new_name in used:
            used[new_name] += 1
            new_name = f"{new_name} #{used[new_name]}"
        else:
            used[new_name] = 1

        rename_map[original_community] = new_name

    df["Community"] = df["OriginalCommunity"].map(rename_map).fillna(df["OriginalCommunity"])
    return df, rename_map


# ============================================================
# WHATSAPP PARSER
# ============================================================

def decode_file(uploaded_file):
    raw = uploaded_file.read()

    for enc in ["utf-8", "utf-8-sig", "utf-16", "latin-1"]:
        try:
            return raw.decode(enc)
        except Exception:
            continue

    return raw.decode("utf-8", errors="ignore")


def parse_datetime(date_str, time_str):
    date_str = str(date_str).strip()
    time_str = str(time_str).strip().replace("\u202f", " ")

    candidates = [
        f"{date_str} {time_str}",
        f"{date_str}, {time_str}",
    ]

    for value in candidates:
        parsed = pd.to_datetime(value, dayfirst=True, errors="coerce")
        if not pd.isna(parsed):
            return parsed.to_pydatetime()

    for value in candidates:
        parsed = pd.to_datetime(value, dayfirst=False, errors="coerce")
        if not pd.isna(parsed):
            return parsed.to_pydatetime()

    return None


def split_sender_message(body):
    """
    Handles both real messages and WhatsApp system lines.

    Examples:
    Vibha Tetr: <Media omitted> -> sender Vibha Tetr
    +91 99999 99999 joined using a group link. -> sender phone, system
    Vibha Tetr pinned a message -> sender Vibha Tetr, system
    Akanksha TETR created group "..." -> sender Akanksha TETR, system
    """
    body = safe_str(body)

    if ": " in body:
        sender, message = body.split(": ", 1)
        return sender.strip(), message.strip(), False

    system_patterns = [
        r"^(?P<actor>.+?) joined using a group link\.?$",
        r"^(?P<actor>.+?) joined using this group's invite link\.?$",
        r"^(?P<actor>.+?) joined using this community's invite link\.?$",
        r"^(?P<actor>.+?) pinned a message\.?$",
        r"^(?P<actor>.+?) changed this group's settings.*$",
        r"^(?P<actor>.+?) changed the group description.*$",
        r"^(?P<actor>.+?) changed the group name.*$",
        r"^(?P<actor>.+?) created group .*$",
        r"^(?P<actor>.+?) added .*$",
        r"^(?P<actor>.+?) was added.*$",
        r"^(?P<actor>.+?) was removed.*$",
        r"^(?P<actor>.+?) left\.?$",
    ]

    for pattern in system_patterns:
        m = re.match(pattern, body, flags=re.I)
        if m:
            actor = m.group("actor").strip()
            return actor, body, True

    return "System", body, True


def parse_whatsapp_export(text, community_name):
    """
    Supports Android and iPhone styles.

    Android:
    15/04/26, 21:31 - Vibha Tetr: Message

    iPhone:
    [15/04/26, 21:31:00] Vibha Tetr: Message
    """
    patterns = [
        re.compile(
            r"^(?P<date>\d{1,2}[/-]\d{1,2}[/-]\d{2,4}),?\s+"
            r"(?P<time>\d{1,2}:\d{2}(?::\d{2})?\s?(?:AM|PM|am|pm)?)\s-\s"
            r"(?P<body>.*)$"
        ),
        re.compile(
            r"^\[(?P<date>\d{1,2}[/-]\d{1,2}[/-]\d{2,4}),?\s+"
            r"(?P<time>\d{1,2}:\d{2}(?::\d{2})?\s?(?:AM|PM|am|pm)?)\]\s"
            r"(?P<body>.*)$"
        )
    ]

    rows = []
    current = None

    for line in text.splitlines():
        line = line.strip("\ufeff").rstrip()
        matched = None

        for pattern in patterns:
            m = pattern.match(line)
            if m:
                matched = m
                break

        if matched:
            if current is not None:
                rows.append(current)

            dt = parse_datetime(matched.group("date"), matched.group("time"))
            body = matched.group("body").strip()
            sender, message, is_system = split_sender_message(body)

            current = {
                "OriginalCommunity": community_name,
                "Community": community_name,
                "DateTime": dt,
                "Sender": sender,
                "Message": message,
                "IsSystem": is_system,
                "RawLine": line
            }
        else:
            if current is not None:
                current["Message"] += "\n" + line
            elif line:
                rows.append({
                    "OriginalCommunity": community_name,
                    "Community": community_name,
                    "DateTime": None,
                    "Sender": "Unknown",
                    "Message": line,
                    "IsSystem": True,
                    "RawLine": line
                })

    if current is not None:
        rows.append(current)

    df = pd.DataFrame(rows)

    if df.empty:
        return df

    df = df.dropna(subset=["DateTime"]).copy()

    if df.empty:
        return df

    df["DateTime"] = pd.to_datetime(df["DateTime"])
    df["Date"] = df["DateTime"].dt.date
    df["Time"] = df["DateTime"].dt.strftime("%H:%M")
    df["Hour"] = df["DateTime"].dt.hour
    df["Message"] = df["Message"].fillna("").astype(str)
    df["Sender"] = df["Sender"].fillna("Unknown").astype(str)

    return df


# ============================================================
# CLASSIFICATION
# ============================================================

def contains_any(text, keywords):
    text = str(text).lower()
    return any(k.lower() in text for k in keywords)


def is_question(message):
    text = str(message).strip().lower()

    if "?" in text:
        return True

    return any(re.search(rf"\b{re.escape(word)}\b", text) for word in QUESTION_WORDS)


def detect_topics(message):
    text = str(message).lower()
    topics = []

    for topic, keywords in TOPIC_KEYWORDS.items():
        if any(keyword.lower() in text for keyword in keywords):
            topics.append(topic)

    return topics if topics else ["General"]


def sentiment_label(message):
    text = str(message).lower()

    neg = sum(1 for k in NEGATIVE_KEYWORDS if k in text)
    pos = sum(1 for k in POSITIVE_KEYWORDS if k in text)

    if neg > pos:
        return "Negative"

    if pos > neg:
        return "Positive"

    return "Neutral"


def has_link(message):
    return bool(re.search(r"https?://|www\.|chat\.whatsapp\.com|forms\.gle|zoom\.us|youtube\.com|youtu\.be", str(message).lower()))


def is_media(message):
    text = str(message).lower()
    markers = [
        "<media omitted>", "image omitted", "video omitted", "audio omitted",
        "sticker omitted", "gif omitted", "document omitted", "voice message omitted"
    ]
    return any(marker in text for marker in markers)


def is_deleted(message):
    text = str(message).lower()
    return "this message was deleted" in text or "you deleted this message" in text


def is_join_event(message):
    text = str(message).lower()

    return (
        "joined using a group link" in text
        or "joined using this group's invite link" in text
        or "joined using this community's invite link" in text
        or "joined a group via invite" in text
        or "you joined a group via invite" in text
        or "was added" in text
        or " added " in f" {text} "
        or "created group" in text
    )


def is_left_event(message):
    text = str(message).lower()
    return (
        " left" in text
        or "was removed" in text
        or " removed " in f" {text} "
    )


def is_pinned_event(message):
    text = str(message).lower()
    return "pinned a message" in text


def is_group_setting_event(message):
    text = str(message).lower()
    return (
        "changed this group's settings" in text
        or "changed the group settings" in text
        or "changed the group description" in text
        or "changed the group name" in text
    )


def is_poll(message):
    text = str(message).lower()
    return "poll:" in text or "\noption:" in text or "option:" in text


def is_announcement_message(message):
    return contains_any(message, ANNOUNCEMENT_KEYWORDS)


def extract_features(df):
    df = df.copy()

    df["IsQuestion"] = df["Message"].apply(is_question)
    df["HasLink"] = df["Message"].apply(has_link)
    df["IsMedia"] = df["Message"].apply(is_media)
    df["IsDeleted"] = df["Message"].apply(is_deleted)
    df["IsJoinEvent"] = df["Message"].apply(is_join_event)
    df["IsLeftEvent"] = df["Message"].apply(is_left_event)
    df["IsPinnedEvent"] = df["Message"].apply(is_pinned_event)
    df["IsGroupSettingEvent"] = df["Message"].apply(is_group_setting_event)
    df["IsPoll"] = df["Message"].apply(is_poll)
    df["IsAnnouncement"] = df["Message"].apply(is_announcement_message)
    df["IsHighIntent"] = df["Message"].apply(lambda x: contains_any(x, HIGH_INTENT_KEYWORDS))
    df["Sentiment"] = df["Message"].apply(sentiment_label)
    df["Topics"] = df["Message"].apply(detect_topics)
    df["TopicPrimary"] = df["Topics"].apply(lambda x: x[0] if x else "General")
    df["MessageLength"] = df["Message"].str.len()

    return df


def mark_answered_questions(df, admin_names=None, response_window_hours=4):
    df = df.copy()

    df["QuestionAnswered"] = False
    df["AnsweredBy"] = ""
    df["NeedsFollowUp"] = False

    admin_names = [a.strip().lower() for a in admin_names or [] if a.strip()]
    question_indices = df.index[df["IsQuestion"] & (~df["IsSystem"])].tolist()

    for idx in question_indices:
        row = df.loc[idx]
        start_time = row["DateTime"]
        end_time = start_time + pd.Timedelta(hours=response_window_hours)

        future = df[
            (df["Community"] == row["Community"])
            & (df["DateTime"] > start_time)
            & (df["DateTime"] <= end_time)
            & (~df["IsSystem"])
            & (df.index != idx)
        ].copy()

        if future.empty:
            df.at[idx, "NeedsFollowUp"] = True
            continue

        if admin_names:
            future["DisplayLower"] = future["DisplayName"].astype(str).str.lower().str.strip()
            future["SenderLower"] = future["Sender"].astype(str).str.lower().str.strip()

            admin_reply = future[
                future["DisplayLower"].isin(admin_names)
                | future["SenderLower"].isin(admin_names)
            ]

            if not admin_reply.empty:
                first_reply = admin_reply.iloc[0]
                df.at[idx, "QuestionAnswered"] = True
                df.at[idx, "AnsweredBy"] = first_reply["DisplayName"]
            else:
                df.at[idx, "NeedsFollowUp"] = True

        else:
            future = future[future["Sender"] != row["Sender"]]

            if not future.empty:
                first_reply = future.iloc[0]
                df.at[idx, "QuestionAnswered"] = True
                df.at[idx, "AnsweredBy"] = first_reply["DisplayName"]
            else:
                df.at[idx, "NeedsFollowUp"] = True

    return df


# ============================================================
# METRICS
# ============================================================

def health_label(score):
    if score >= 75:
        return "Strong"
    if score >= 55:
        return "Moderate"
    if score >= 35:
        return "Low"
    return "Risk"


def calculate_health_score(total_messages, active_members, unanswered, negative_cases, high_intent, questions):
    score = 50

    if total_messages >= 100:
        score += 15
    elif total_messages >= 50:
        score += 10
    elif total_messages >= 20:
        score += 5

    if active_members >= 25:
        score += 15
    elif active_members >= 10:
        score += 10
    elif active_members >= 5:
        score += 5

    if questions > 0:
        unanswered_rate = unanswered / questions
        if unanswered_rate <= 0.1:
            score += 10
        elif unanswered_rate <= 0.3:
            score += 5
        else:
            score -= 10

    if high_intent >= 10:
        score += 10
    elif high_intent >= 3:
        score += 5

    if negative_cases >= 10:
        score -= 15
    elif negative_cases >= 3:
        score -= 8

    return max(0, min(100, score))


def build_daily_community_metrics(df):
    if df.empty:
        return pd.DataFrame()

    records = []

    for (dt, community), g in df.groupby(["Date", "Community"], dropna=False):
        non_system = g[~g["IsSystem"]]
        questions = g[g["IsQuestion"] & (~g["IsSystem"])]
        answered = questions[questions["QuestionAnswered"]]
        unanswered = questions[~questions["QuestionAnswered"]]

        sentiment_counts = non_system["Sentiment"].value_counts().to_dict()
        topic_counts = non_system["TopicPrimary"].value_counts()
        top_topic = topic_counts.index[0] if not topic_counts.empty else "General"

        active_members = non_system["DisplayName"].nunique()
        hour_counts = non_system["Hour"].value_counts()
        peak_hour = int(hour_counts.idxmax()) if not hour_counts.empty else ""

        negative_cases = non_system[non_system["Sentiment"] == "Negative"].shape[0]
        high_intent = non_system[non_system["IsHighIntent"]].shape[0]
        total_messages = len(non_system)

        health_score = calculate_health_score(
            total_messages=total_messages,
            active_members=active_members,
            unanswered=len(unanswered),
            negative_cases=negative_cases,
            high_intent=high_intent,
            questions=len(questions)
        )

        records.append({
            "Date": dt,
            "Community": community,
            "Total Messages": total_messages,
            "Active Members": active_members,
            "Matched Messages": int(non_system["IsMatched"].sum()),
            "Student Messages": int((non_system["PersonType"] == "Student").sum()),
            "Persona Messages": int((non_system["PersonType"] == "Persona").sum()),
            "Unmatched Messages": int((non_system["PersonType"] == "Unmatched").sum()),
            "Unique Matched People": int(non_system[non_system["IsMatched"]]["DisplayName"].nunique()),
            "Unique Students": int(non_system[non_system["PersonType"] == "Student"]["DisplayName"].nunique()),
            "Unique Personas": int(non_system[non_system["PersonType"] == "Persona"]["DisplayName"].nunique()),
            "System Messages": int(g["IsSystem"].sum()),
            "New Member Events": int(g["IsJoinEvent"].sum()),
            "Left / Removed Events": int(g["IsLeftEvent"].sum()),
            "Pinned Messages": int(g["IsPinnedEvent"].sum()),
            "Group Setting Events": int(g["IsGroupSettingEvent"].sum()),
            "Polls Shared": int(non_system["IsPoll"].sum()),
            "Announcements": int(non_system["IsAnnouncement"].sum()),
            "Media Shared": int(g["IsMedia"].sum()),
            "Links Shared": int(g["HasLink"].sum()),
            "Deleted Messages": int(g["IsDeleted"].sum()),
            "Questions Asked": len(questions),
            "Questions Answered": len(answered),
            "Unanswered Questions": len(unanswered),
            "High Intent Messages": high_intent,
            "Negative Cases": negative_cases,
            "Positive Messages": sentiment_counts.get("Positive", 0),
            "Neutral Messages": sentiment_counts.get("Neutral", 0),
            "Negative Messages": sentiment_counts.get("Negative", 0),
            "Peak Activity Hour": peak_hour,
            "Main Topic": top_topic,
            "Community Health Score": health_score,
            "Overall Health": health_label(health_score)
        })

    return pd.DataFrame(records).sort_values(["Date", "Community"])


def build_topic_summary(df):
    if df.empty:
        return pd.DataFrame()

    exploded = df[~df["IsSystem"]].explode("Topics")

    return (
        exploded
        .groupby(["Date", "Community", "Topics"])
        .size()
        .reset_index(name="Message Count")
        .rename(columns={"Topics": "Topic"})
        .sort_values(["Date", "Community", "Message Count"], ascending=[True, True, False])
    )


def build_top_members(df):
    if df.empty:
        return pd.DataFrame()

    non_system = df[~df["IsSystem"]]

    return (
        non_system
        .groupby([
            "Date", "Community", "DisplayName", "Sender", "SenderPhone",
            "MatchedName", "PersonType", "BatchLabel"
        ], dropna=False)
        .size()
        .reset_index(name="Messages")
        .sort_values(["Date", "Community", "Messages"], ascending=[True, True, False])
    )


def build_hourly_summary(df):
    if df.empty:
        return pd.DataFrame()

    return (
        df[~df["IsSystem"]]
        .groupby(["Date", "Community", "Hour"])
        .size()
        .reset_index(name="Messages")
        .sort_values(["Date", "Community", "Hour"])
    )


def get_questions_tracker(df):
    q = df[(~df["IsSystem"]) & (df["IsQuestion"])].copy()

    cols = [
        "Date", "Time", "Community", "DisplayName", "PersonType", "BatchLabel",
        "Sender", "SenderPhone", "MatchedName", "Message",
        "QuestionAnswered", "AnsweredBy", "NeedsFollowUp"
    ]

    if q.empty:
        return pd.DataFrame(columns=cols)

    return q[cols].sort_values(["Date", "Time", "Community"])


def get_high_intent_tracker(df):
    h = df[(~df["IsSystem"]) & (df["IsHighIntent"])].copy()

    cols = [
        "Date", "Time", "Community", "DisplayName", "PersonType", "BatchLabel",
        "Sender", "SenderPhone", "MatchedName", "Message", "Sentiment"
    ]

    if h.empty:
        return pd.DataFrame(columns=cols)

    return h[cols].sort_values(["Date", "Time", "Community"])


def get_negative_tracker(df):
    n = df[(~df["IsSystem"]) & (df["Sentiment"] == "Negative")].copy()

    cols = [
        "Date", "Time", "Community", "DisplayName", "PersonType", "BatchLabel",
        "Sender", "SenderPhone", "MatchedName", "Message"
    ]

    if n.empty:
        return pd.DataFrame(columns=cols)

    return n[cols].sort_values(["Date", "Time", "Community"])


def get_announcements(df):
    a = df[(~df["IsSystem"]) & (df["IsAnnouncement"])].copy()

    cols = [
        "Date", "Time", "Community", "DisplayName", "PersonType",
        "BatchLabel", "Sender", "Message"
    ]

    if a.empty:
        return pd.DataFrame(columns=cols)

    return a[cols].sort_values(["Date", "Time", "Community"])


def get_polls(df):
    p = df[(~df["IsSystem"]) & (df["IsPoll"])].copy()

    cols = [
        "Date", "Time", "Community", "DisplayName", "PersonType",
        "BatchLabel", "Sender", "Message"
    ]

    if p.empty:
        return pd.DataFrame(columns=cols)

    return p[cols].sort_values(["Date", "Time", "Community"])


def get_system_events(df):
    s = df[df["IsSystem"]].copy()

    cols = [
        "Date", "Time", "Community", "DisplayName", "PersonType",
        "BatchLabel", "Sender", "Message",
        "IsJoinEvent", "IsLeftEvent", "IsPinnedEvent", "IsGroupSettingEvent"
    ]

    available_cols = [c for c in cols if c in s.columns]

    if s.empty:
        return pd.DataFrame(columns=cols)

    return s[available_cols].sort_values(["Date", "Time", "Community"])


def get_important_messages(df, limit=300):
    if df.empty:
        return pd.DataFrame()

    important = df[
        (~df["IsSystem"])
        & (
            df["IsQuestion"]
            | df["IsHighIntent"]
            | (df["Sentiment"] == "Negative")
            | df["IsAnnouncement"]
            | df["IsPoll"]
        )
    ].copy()

    cols = [
        "Date", "Time", "Community", "DisplayName", "PersonType", "BatchLabel",
        "Sender", "SenderPhone", "MatchedName", "Category", "Sentiment",
        "QuestionAnswered", "NeedsFollowUp", "Message"
    ]

    if important.empty:
        return pd.DataFrame(columns=cols)

    important["Category"] = np.select(
        [
            important["IsQuestion"],
            important["IsPoll"],
            important["IsAnnouncement"],
            important["IsHighIntent"],
            important["Sentiment"].eq("Negative")
        ],
        [
            "Question",
            "Poll",
            "Announcement",
            "High Intent",
            "Negative / Risk"
        ],
        default="Important"
    )

    return important[cols].sort_values(["Date", "Time"]).head(limit)


def detect_admin_announcements(df, admin_names=None):
    admin_names = [a.strip().lower() for a in admin_names or [] if a.strip()]
    data = df[~df["IsSystem"]].copy()

    if admin_names:
        data["DisplayLower"] = data["DisplayName"].astype(str).str.lower().str.strip()
        data["SenderLower"] = data["Sender"].astype(str).str.lower().str.strip()

        data = data[
            data["DisplayLower"].isin(admin_names)
            | data["SenderLower"].isin(admin_names)
        ]

    data = data[data["Message"].apply(lambda x: contains_any(x, ANNOUNCEMENT_KEYWORDS))]

    cols = ["Date", "Time", "Community", "DisplayName", "PersonType", "BatchLabel", "Sender", "Message"]

    if data.empty:
        return pd.DataFrame(columns=cols)

    return data[cols].sort_values(["Date", "Time", "Community"])


# ============================================================
# STYLING
# ============================================================

def style_personas(df):
    if df is None or df.empty:
        return df

    def row_style(row):
        person_type = str(row.get("PersonType", "")).lower()

        if person_type == "persona":
            return ["background-color: #FFF3CD; color: #5F3B00; font-weight: 600;" for _ in row]

        return ["" for _ in row]

    return df.style.apply(row_style, axis=1)


# ============================================================
# EXPORTS
# ============================================================

def clean_for_excel(df):
    if df is None or df.empty:
        return pd.DataFrame()

    out = df.copy()

    for col in out.columns:
        if out[col].dtype == "object":
            out[col] = out[col].astype(str)

    return out


def build_excel_report(
    metrics_df,
    topic_summary,
    top_members,
    questions_df,
    high_intent_df,
    negative_df,
    important_df,
    announcements_df,
    polls_df,
    system_events_df,
    raw_df
):
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        clean_for_excel(metrics_df).to_excel(writer, sheet_name="Daily Metrics", index=False)
        clean_for_excel(topic_summary).to_excel(writer, sheet_name="Topic Summary", index=False)
        clean_for_excel(top_members).to_excel(writer, sheet_name="Top Members", index=False)
        clean_for_excel(questions_df).to_excel(writer, sheet_name="Questions", index=False)
        clean_for_excel(high_intent_df).to_excel(writer, sheet_name="High Intent", index=False)
        clean_for_excel(negative_df).to_excel(writer, sheet_name="Negative Cases", index=False)
        clean_for_excel(announcements_df).to_excel(writer, sheet_name="Announcements", index=False)
        clean_for_excel(polls_df).to_excel(writer, sheet_name="Polls", index=False)
        clean_for_excel(system_events_df).to_excel(writer, sheet_name="System Events", index=False)
        clean_for_excel(important_df).to_excel(writer, sheet_name="Important Messages", index=False)

        export_raw = raw_df.copy()
        export_raw["Topics"] = export_raw["Topics"].astype(str)
        clean_for_excel(export_raw).to_excel(writer, sheet_name="Parsed Raw Data", index=False)

        wb = writer.book

        header_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
        persona_fill = PatternFill(fill_type="solid", fgColor="FFF3CD")
        header_font = Font(bold=True)
        persona_font = Font(bold=True, color="5F3B00")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.freeze_panes = "A2"

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            headers = [cell.value for cell in ws[1]]
            person_type_col = None

            if "PersonType" in headers:
                person_type_col = headers.index("PersonType") + 1

            if person_type_col:
                for row in range(2, ws.max_row + 1):
                    if str(ws.cell(row=row, column=person_type_col).value).lower() == "persona":
                        for col in range(1, ws.max_column + 1):
                            ws.cell(row=row, column=col).fill = persona_fill
                            ws.cell(row=row, column=col).font = persona_font

            for column_cells in ws.columns:
                max_len = 0
                col_letter = column_cells[0].column_letter

                for cell in column_cells[:120]:
                    try:
                        max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                    except Exception:
                        pass

                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 48)

    output.seek(0)
    return output


def df_to_html_table(df, title):
    if df is None or df.empty:
        return f"""
        <section>
            <h2>{html.escape(title)}</h2>
            <p class="muted">No data available.</p>
        </section>
        """

    safe_df = df.copy()

    for col in safe_df.columns:
        safe_df[col] = safe_df[col].astype(str).map(html.escape)

    headers = "".join(f"<th>{html.escape(str(c))}</th>" for c in safe_df.columns)
    rows = ""

    for _, row in safe_df.iterrows():
        is_persona = str(row.get("PersonType", "")).lower() == "persona"
        cls = "persona-row" if is_persona else ""
        cells = "".join(f"<td>{row[c]}</td>" for c in safe_df.columns)
        rows += f"<tr class='{cls}'>{cells}</tr>"

    return f"""
    <section>
        <h2>{html.escape(title)}</h2>
        <table class="report-table">
            <thead><tr>{headers}</tr></thead>
            <tbody>{rows}</tbody>
        </table>
    </section>
    """


def build_html_report(
    title,
    filtered_df,
    metrics_df,
    topic_summary,
    top_members,
    questions_df,
    high_intent_df,
    negative_df,
    important_df,
    announcements_df,
    polls_df,
    system_events_df,
    start_date,
    end_date
):
    generated_at = datetime.now().strftime("%d %b %Y, %I:%M %p")

    total_messages = int(metrics_df["Total Messages"].sum()) if not metrics_df.empty else 0
    active_members = filtered_df[~filtered_df["IsSystem"]]["DisplayName"].nunique() if not filtered_df.empty else 0
    matched_people = filtered_df[(~filtered_df["IsSystem"]) & (filtered_df["IsMatched"])]["DisplayName"].nunique() if not filtered_df.empty else 0
    personas = filtered_df[(~filtered_df["IsSystem"]) & (filtered_df["PersonType"] == "Persona")]["DisplayName"].nunique() if not filtered_df.empty else 0
    total_questions = int(metrics_df["Questions Asked"].sum()) if not metrics_df.empty else 0
    unanswered = int(metrics_df["Unanswered Questions"].sum()) if not metrics_df.empty else 0
    high_intent = int(metrics_df["High Intent Messages"].sum()) if not metrics_df.empty else 0
    negative = int(metrics_df["Negative Cases"].sum()) if not metrics_df.empty else 0

    return f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="utf-8">
        <title>{html.escape(title)}</title>
        <style>
            body {{
                font-family: Arial, Helvetica, sans-serif;
                margin: 0;
                padding: 0;
                background: #f5f7fb;
                color: #172033;
            }}
            .container {{
                max-width: 1180px;
                margin: 0 auto;
                padding: 32px;
            }}
            .header {{
                background: linear-gradient(135deg, #172033, #31415f);
                color: white;
                padding: 32px;
                border-radius: 18px;
                margin-bottom: 24px;
            }}
            .header h1 {{
                margin: 0 0 10px 0;
                font-size: 30px;
            }}
            .header p {{
                margin: 4px 0;
                opacity: 0.9;
            }}
            .kpi-grid {{
                display: grid;
                grid-template-columns: repeat(8, 1fr);
                gap: 14px;
                margin-bottom: 26px;
            }}
            .kpi {{
                background: white;
                border-radius: 16px;
                padding: 18px;
                box-shadow: 0 8px 24px rgba(23, 32, 51, 0.08);
            }}
            .kpi .label {{
                color: #667085;
                font-size: 12px;
                text-transform: uppercase;
                letter-spacing: 0.04em;
            }}
            .kpi .value {{
                font-size: 25px;
                font-weight: 700;
                margin-top: 8px;
            }}
            section {{
                background: white;
                padding: 24px;
                border-radius: 18px;
                margin-bottom: 22px;
                box-shadow: 0 8px 24px rgba(23, 32, 51, 0.08);
            }}
            h2 {{
                margin-top: 0;
                color: #172033;
                font-size: 22px;
                border-bottom: 1px solid #e7eaf0;
                padding-bottom: 10px;
            }}
            .report-table {{
                width: 100%;
                border-collapse: collapse;
                font-size: 13px;
            }}
            .report-table th {{
                background: #eef2f7;
                color: #172033;
                text-align: left;
                padding: 10px;
                border: 1px solid #d9dee8;
            }}
            .report-table td {{
                padding: 9px 10px;
                border: 1px solid #e4e7ee;
                vertical-align: top;
            }}
            .report-table tr:nth-child(even) {{
                background: #fafbfc;
            }}
            .persona-row td {{
                background: #fff3cd !important;
                color: #5f3b00;
                font-weight: 600;
            }}
            .muted {{
                color: #667085;
            }}
            .footer {{
                text-align: center;
                color: #667085;
                margin-top: 28px;
                font-size: 12px;
            }}
            @media print {{
                body {{
                    background: white;
                }}
                .container {{
                    padding: 0;
                }}
                section, .kpi {{
                    box-shadow: none;
                    border: 1px solid #e4e7ee;
                }}
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>{html.escape(title)}</h1>
                <p><strong>Reporting Period:</strong> {start_date.strftime("%d %b %Y")} to {end_date.strftime("%d %b %Y")}</p>
                <p><strong>Generated:</strong> {generated_at}</p>
                <p><strong>Persona Highlight:</strong> Persona rows are highlighted in yellow.</p>
            </div>

            <div class="kpi-grid">
                <div class="kpi"><div class="label">Messages</div><div class="value">{total_messages}</div></div>
                <div class="kpi"><div class="label">Active Members</div><div class="value">{active_members}</div></div>
                <div class="kpi"><div class="label">Matched People</div><div class="value">{matched_people}</div></div>
                <div class="kpi"><div class="label">Personas</div><div class="value">{personas}</div></div>
                <div class="kpi"><div class="label">Questions</div><div class="value">{total_questions}</div></div>
                <div class="kpi"><div class="label">Unanswered</div><div class="value">{unanswered}</div></div>
                <div class="kpi"><div class="label">High Intent</div><div class="value">{high_intent}</div></div>
                <div class="kpi"><div class="label">Negative Cases</div><div class="value">{negative}</div></div>
            </div>

            <section>
                <h2>Executive Summary</h2>
                <p>
                    This report summarises WhatsApp community activity for the selected period.
                    Students and personas are matched from the workbook. Persona rows are highlighted separately.
                    Polls, pinned messages, links, media, system events, questions, announcements and risk signals are tracked.
                </p>
            </section>

            {df_to_html_table(metrics_df, "Daily Performance")}
            {df_to_html_table(topic_summary, "Topic Summary")}
            {df_to_html_table(top_members.head(100), "Top Active Members")}
            {df_to_html_table(questions_df, "Questions Tracker")}
            {df_to_html_table(high_intent_df, "High Intent / Lead Signals")}
            {df_to_html_table(negative_df, "Negative Sentiment / Risk Cases")}
            {df_to_html_table(announcements_df, "Announcements")}
            {df_to_html_table(polls_df, "Polls")}
            {df_to_html_table(system_events_df, "System Events")}
            {df_to_html_table(important_df, "Important Conversations")}

            <div class="footer">
                Generated automatically from WhatsApp chat exports.
            </div>
        </div>
    </body>
    </html>
    """


# ============================================================
# PREPARE DATA
# ============================================================

def prepare_report_data(df):
    return {
        "metrics_df": build_daily_community_metrics(df),
        "topic_summary": build_topic_summary(df),
        "top_members": build_top_members(df),
        "hourly_summary": build_hourly_summary(df),
        "questions_df": get_questions_tracker(df),
        "high_intent_df": get_high_intent_tracker(df),
        "negative_df": get_negative_tracker(df),
        "announcements_df": get_announcements(df),
        "polls_df": get_polls(df),
        "system_events_df": get_system_events(df),
        "important_df": get_important_messages(df),
    }


# ============================================================
# UI RENDERING
# ============================================================

def render_kpis(metrics_df, filtered_df):
    total_messages = int(metrics_df["Total Messages"].sum()) if not metrics_df.empty else 0
    active_members = filtered_df[~filtered_df["IsSystem"]]["DisplayName"].nunique() if not filtered_df.empty else 0
    matched_people = filtered_df[(~filtered_df["IsSystem"]) & (filtered_df["IsMatched"])]["DisplayName"].nunique() if not filtered_df.empty else 0
    personas = filtered_df[(~filtered_df["IsSystem"]) & (filtered_df["PersonType"] == "Persona")]["DisplayName"].nunique() if not filtered_df.empty else 0
    total_questions = int(metrics_df["Questions Asked"].sum()) if not metrics_df.empty else 0
    polls = int(metrics_df["Polls Shared"].sum()) if not metrics_df.empty and "Polls Shared" in metrics_df.columns else 0
    negative = int(metrics_df["Negative Cases"].sum()) if not metrics_df.empty else 0

    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
    c1.metric("Messages", total_messages)
    c2.metric("Active Members", active_members)
    c3.metric("Matched People", matched_people)
    c4.metric("Personas", personas)
    c5.metric("Questions", total_questions)
    c6.metric("Polls", polls)
    c7.metric("Negative Cases", negative)


def render_charts(filtered_df, metrics_df, topic_summary, hourly_summary, title_prefix=""):
    st.subheader("Visual Analytics")

    col1, col2 = st.columns(2)

    with col1:
        if not metrics_df.empty:
            daily = metrics_df.groupby("Date")["Total Messages"].sum().reset_index()
            fig = px.bar(
                daily,
                x="Date",
                y="Total Messages",
                title=f"{title_prefix}Daily Messages"
            )
            st.plotly_chart(fig, use_container_width=True)

    with col2:
        if not topic_summary.empty:
            topic_total = (
                topic_summary
                .groupby("Topic")["Message Count"]
                .sum()
                .reset_index()
                .sort_values("Message Count", ascending=False)
            )
            fig = px.pie(
                topic_total,
                names="Topic",
                values="Message Count",
                title=f"{title_prefix}Topic Distribution"
            )
            st.plotly_chart(fig, use_container_width=True)

    col3, col4 = st.columns(2)

    with col3:
        if not filtered_df.empty:
            people_type = (
                filtered_df[~filtered_df["IsSystem"]]
                .groupby("PersonType")
                .size()
                .reset_index(name="Messages")
            )
            fig = px.bar(
                people_type,
                x="PersonType",
                y="Messages",
                title=f"{title_prefix}Student / Persona / Unmatched Split"
            )
            st.plotly_chart(fig, use_container_width=True)

    with col4:
        if not hourly_summary.empty:
            hourly_total = hourly_summary.groupby("Hour")["Messages"].sum().reset_index()
            fig = px.line(
                hourly_total,
                x="Hour",
                y="Messages",
                markers=True,
                title=f"{title_prefix}Hourly Activity"
            )
            st.plotly_chart(fig, use_container_width=True)


def render_downloads(
    title,
    df,
    metrics_df,
    topic_summary,
    top_members,
    questions_df,
    high_intent_df,
    negative_df,
    important_df,
    announcements_df,
    polls_df,
    system_events_df,
    start_date,
    end_date,
    key_prefix
):
    html_report = build_html_report(
        title=title,
        filtered_df=df,
        metrics_df=metrics_df,
        topic_summary=topic_summary,
        top_members=top_members,
        questions_df=questions_df,
        high_intent_df=high_intent_df,
        negative_df=negative_df,
        important_df=important_df,
        announcements_df=announcements_df,
        polls_df=polls_df,
        system_events_df=system_events_df,
        start_date=start_date,
        end_date=end_date
    )

    excel_report = build_excel_report(
        metrics_df=metrics_df,
        topic_summary=topic_summary,
        top_members=top_members,
        questions_df=questions_df,
        high_intent_df=high_intent_df,
        negative_df=negative_df,
        important_df=important_df,
        announcements_df=announcements_df,
        polls_df=polls_df,
        system_events_df=system_events_df,
        raw_df=df
    )

    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", title).strip("_")
    report_name_base = f"{safe_name}_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}"

    c1, c2 = st.columns(2)

    with c1:
        st.download_button(
            label="Download HTML Report",
            data=html_report.encode("utf-8"),
            file_name=f"{report_name_base}.html",
            mime="text/html",
            use_container_width=True,
            key=f"{key_prefix}_html"
        )

    with c2:
        st.download_button(
            label="Download Excel Report",
            data=excel_report,
            file_name=f"{report_name_base}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"{key_prefix}_excel"
        )


def render_community_page(community_name, community_df, admin_names, start_date, end_date):
    st.header(f"📌 {community_name}")

    data = prepare_report_data(community_df)

    metrics_df = data["metrics_df"]
    topic_summary = data["topic_summary"]
    top_members = data["top_members"]
    hourly_summary = data["hourly_summary"]
    questions_df = data["questions_df"]
    high_intent_df = data["high_intent_df"]
    negative_df = data["negative_df"]
    announcements_df = data["announcements_df"]
    polls_df = data["polls_df"]
    system_events_df = data["system_events_df"]
    important_df = data["important_df"]

    render_kpis(metrics_df, community_df)
    st.markdown("---")

    day_count = (end_date - start_date).days + 1

    if day_count == 2:
        st.subheader("Two-Day Comparison")
        daily = metrics_df.groupby("Date").sum(numeric_only=True).reset_index()

        if daily["Date"].nunique() == 2:
            d1, d2 = sorted(daily["Date"].unique())
            row1 = daily[daily["Date"] == d1].iloc[0]
            row2 = daily[daily["Date"] == d2].iloc[0]

            comparison_metrics = [
                "Total Messages", "Active Members", "Matched Messages",
                "Student Messages", "Persona Messages", "Unmatched Messages",
                "Questions Asked", "Unanswered Questions", "High Intent Messages",
                "Negative Cases", "Polls Shared", "Media Shared", "Links Shared",
                "Pinned Messages", "New Member Events"
            ]

            rows = []

            for metric in comparison_metrics:
                v1 = int(row1.get(metric, 0))
                v2 = int(row2.get(metric, 0))
                change = v2 - v1
                pct_change = ""

                if v1 != 0:
                    pct_change = f"{((v2 - v1) / v1) * 100:.1f}%"

                rows.append({
                    "Metric": metric,
                    d1.strftime("%d %b %Y"): v1,
                    d2.strftime("%d %b %Y"): v2,
                    "Change": change,
                    "% Change": pct_change
                })

            st.dataframe(pd.DataFrame(rows), use_container_width=True)
        else:
            st.info("Only one selected day has data for this community.")

    elif day_count >= 3:
        st.subheader("Date-Wise Trend")
        daily = metrics_df.groupby("Date").sum(numeric_only=True).reset_index()

        metric_options = [
            "Total Messages", "Active Members", "Matched Messages",
            "Student Messages", "Persona Messages", "Unmatched Messages",
            "Questions Asked", "Unanswered Questions", "High Intent Messages",
            "Negative Cases", "Polls Shared", "Media Shared", "Links Shared"
        ]

        selected_metric = st.selectbox(
            f"Select trend metric for {community_name}",
            metric_options,
            key=f"trend_{community_name}"
        )

        if selected_metric in daily.columns:
            fig = px.line(
                daily,
                x="Date",
                y=selected_metric,
                markers=True,
                title=f"{community_name} - {selected_metric}"
            )
            st.plotly_chart(fig, use_container_width=True)

        st.dataframe(daily, use_container_width=True)
    else:
        st.subheader("Single-Day Data")
        st.dataframe(metrics_df, use_container_width=True)

    render_charts(community_df, metrics_df, topic_summary, hourly_summary, title_prefix=f"{community_name} - ")

    st.markdown("---")

    tabs = st.tabs([
        "Daily Metrics",
        "Topics",
        "Top Members",
        "Matched People",
        "Questions",
        "High Intent",
        "Negative Cases",
        "Announcements",
        "Polls",
        "System Events",
        "Important Conversations",
        "Raw Data"
    ])

    with tabs[0]:
        st.dataframe(metrics_df, use_container_width=True)

    with tabs[1]:
        st.dataframe(topic_summary, use_container_width=True)

    with tabs[2]:
        st.dataframe(style_personas(top_members), use_container_width=True)

    with tabs[3]:
        matched = community_df[
            (~community_df["IsSystem"])
            & (community_df["IsMatched"])
        ][[
            "Date", "Time", "Community", "DisplayName", "PersonType", "BatchLabel",
            "Sender", "SenderPhone", "MatchedName", "Message"
        ]].copy()

        unmatched = community_df[
            (~community_df["IsSystem"])
            & (~community_df["IsMatched"])
        ][[
            "Date", "Time", "Community", "Sender", "SenderPhone", "Message"
        ]].copy()

        c1, c2, c3 = st.columns(3)
        c1.metric("Matched Messages", len(matched))
        c2.metric("Persona Messages", int((matched["PersonType"] == "Persona").sum()) if not matched.empty else 0)
        c3.metric("Unmatched Messages", len(unmatched))

        st.markdown("#### Matched Students & Personas")
        st.dataframe(style_personas(matched), use_container_width=True)

        st.markdown("#### Unmatched Senders")
        st.dataframe(unmatched, use_container_width=True)

    with tabs[4]:
        st.dataframe(style_personas(questions_df), use_container_width=True)

        unanswered = questions_df[questions_df["QuestionAnswered"] == False]
        if not unanswered.empty:
            st.warning(f"{len(unanswered)} unanswered question(s) detected.")
            st.dataframe(style_personas(unanswered), use_container_width=True)

    with tabs[5]:
        st.dataframe(style_personas(high_intent_df), use_container_width=True)

    with tabs[6]:
        st.dataframe(style_personas(negative_df), use_container_width=True)

    with tabs[7]:
        st.dataframe(style_personas(announcements_df), use_container_width=True)

    with tabs[8]:
        st.dataframe(style_personas(polls_df), use_container_width=True)

    with tabs[9]:
        st.dataframe(style_personas(system_events_df), use_container_width=True)

    with tabs[10]:
        st.dataframe(style_personas(important_df), use_container_width=True)

    with tabs[11]:
        raw_cols = [
            "DateTime", "Date", "Time", "OriginalCommunity", "Community",
            "Sender", "DisplayName", "PersonType", "BatchLabel",
            "SenderPhone", "MatchedName", "IsMatched", "IsPersona",
            "Message", "IsSystem", "IsQuestion", "QuestionAnswered",
            "NeedsFollowUp", "IsHighIntent", "IsAnnouncement", "IsPoll",
            "IsJoinEvent", "IsPinnedEvent", "Sentiment", "TopicPrimary",
            "HasLink", "IsMedia"
        ]

        available = [c for c in raw_cols if c in community_df.columns]
        st.dataframe(style_personas(community_df[available]), use_container_width=True)

    st.markdown("---")
    st.subheader(f"Download Report - {community_name}")

    render_downloads(
        title=f"{community_name} WhatsApp Community Report",
        df=community_df,
        metrics_df=metrics_df,
        topic_summary=topic_summary,
        top_members=top_members,
        questions_df=questions_df,
        high_intent_df=high_intent_df,
        negative_df=negative_df,
        important_df=important_df,
        announcements_df=announcements_df,
        polls_df=polls_df,
        system_events_df=system_events_df,
        start_date=start_date,
        end_date=end_date,
        key_prefix=f"download_{community_name}"
    )


def render_all_groups_comparison(filtered_df, admin_names, start_date, end_date):
    st.header("📊 All Groups Comparison")

    data = prepare_report_data(filtered_df)

    metrics_df = data["metrics_df"]
    topic_summary = data["topic_summary"]
    top_members = data["top_members"]
    questions_df = data["questions_df"]
    high_intent_df = data["high_intent_df"]
    negative_df = data["negative_df"]
    announcements_df = data["announcements_df"]
    polls_df = data["polls_df"]
    system_events_df = data["system_events_df"]
    important_df = data["important_df"]

    if metrics_df.empty:
        st.info("No comparison data available.")
        return

    comparison = (
        metrics_df
        .groupby("Community")
        .agg({
            "Total Messages": "sum",
            "Active Members": "sum",
            "Matched Messages": "sum",
            "Student Messages": "sum",
            "Persona Messages": "sum",
            "Unmatched Messages": "sum",
            "Unique Matched People": "sum",
            "Unique Students": "sum",
            "Unique Personas": "sum",
            "Questions Asked": "sum",
            "Questions Answered": "sum",
            "Unanswered Questions": "sum",
            "High Intent Messages": "sum",
            "Negative Cases": "sum",
            "Polls Shared": "sum",
            "Announcements": "sum",
            "Media Shared": "sum",
            "Links Shared": "sum",
            "Pinned Messages": "sum",
            "New Member Events": "sum",
            "Community Health Score": "mean"
        })
        .reset_index()
    )

    comparison["Community Health Score"] = comparison["Community Health Score"].round(1)
    comparison["Overall Health"] = comparison["Community Health Score"].apply(health_label)

    st.subheader("Group-Wise Comparison")
    st.dataframe(comparison, use_container_width=True)

    col1, col2 = st.columns(2)

    with col1:
        fig = px.bar(
            comparison.sort_values("Total Messages", ascending=False),
            x="Community",
            y="Total Messages",
            title="Total Messages by Group"
        )
        st.plotly_chart(fig, use_container_width=True)

    with col2:
        fig = px.bar(
            comparison.sort_values("Persona Messages", ascending=False),
            x="Community",
            y="Persona Messages",
            title="Persona Messages by Group"
        )
        st.plotly_chart(fig, use_container_width=True)

    col3, col4 = st.columns(2)

    with col3:
        fig = px.bar(
            comparison.sort_values("High Intent Messages", ascending=False),
            x="Community",
            y="High Intent Messages",
            title="High Intent Messages by Group"
        )
        st.plotly_chart(fig, use_container_width=True)

    with col4:
        fig = px.bar(
            comparison.sort_values("Negative Cases", ascending=False),
            x="Community",
            y="Negative Cases",
            title="Negative Cases by Group"
        )
        st.plotly_chart(fig, use_container_width=True)

    st.markdown("---")
    st.subheader("Date-Wise Group Comparison")

    daily_group = metrics_df.groupby(["Date", "Community"]).sum(numeric_only=True).reset_index()

    metric_options = [
        "Total Messages", "Active Members", "Matched Messages",
        "Student Messages", "Persona Messages", "Unmatched Messages",
        "Questions Asked", "Unanswered Questions", "High Intent Messages",
        "Negative Cases", "Polls Shared", "Announcements", "Media Shared",
        "Links Shared"
    ]

    selected_metric = st.selectbox("Select metric", metric_options, key="all_groups_metric")

    if selected_metric in daily_group.columns:
        fig = px.line(
            daily_group,
            x="Date",
            y=selected_metric,
            color="Community",
            markers=True,
            title=f"Date-Wise Comparison: {selected_metric}"
        )
        st.plotly_chart(fig, use_container_width=True)

    st.dataframe(daily_group, use_container_width=True)

    tabs = st.tabs([
        "Daily Metrics",
        "Topics",
        "Top Members",
        "Questions",
        "High Intent",
        "Negative Cases",
        "Announcements",
        "Polls",
        "System Events",
        "Important Conversations"
    ])

    with tabs[0]:
        st.dataframe(metrics_df, use_container_width=True)

    with tabs[1]:
        st.dataframe(topic_summary, use_container_width=True)

    with tabs[2]:
        st.dataframe(style_personas(top_members), use_container_width=True)

    with tabs[3]:
        st.dataframe(style_personas(questions_df), use_container_width=True)

    with tabs[4]:
        st.dataframe(style_personas(high_intent_df), use_container_width=True)

    with tabs[5]:
        st.dataframe(style_personas(negative_df), use_container_width=True)

    with tabs[6]:
        st.dataframe(style_personas(announcements_df), use_container_width=True)

    with tabs[7]:
        st.dataframe(style_personas(polls_df), use_container_width=True)

    with tabs[8]:
        st.dataframe(style_personas(system_events_df), use_container_width=True)

    with tabs[9]:
        st.dataframe(style_personas(important_df), use_container_width=True)

    st.markdown("---")
    st.subheader("Download All Groups Comparison Report")

    render_downloads(
        title="All Groups Comparison WhatsApp Community Report",
        df=filtered_df,
        metrics_df=metrics_df,
        topic_summary=topic_summary,
        top_members=top_members,
        questions_df=questions_df,
        high_intent_df=high_intent_df,
        negative_df=negative_df,
        important_df=important_df,
        announcements_df=announcements_df,
        polls_df=polls_df,
        system_events_df=system_events_df,
        start_date=start_date,
        end_date=end_date,
        key_prefix="all_groups_download"
    )


# ============================================================
# MAIN
# ============================================================

def main():
    st.title("💬 Tetr Community Analysis")
    st.caption(
        "Upload WhatsApp chat exports, auto-match students/personas from students_phone.xlsx, "
        "rename communities by dominant UG/PG batch, and download professional reports."
    )

    phone_mapping, name_mapping, students_df, personas_df, mapping_error = load_people_mapping()

    with st.sidebar:
        st.header("Upload & Settings")

        uploaded_files = st.file_uploader(
            "Upload WhatsApp chat exports",
            type=["txt"],
            accept_multiple_files=True
        )

        st.markdown("---")

        if mapping_error:
            st.warning(mapping_error)
        else:
            student_rows = len(students_df) if students_df is not None else 0
            persona_rows = len(personas_df) if personas_df is not None else 0

            st.success(
                f"Loaded phone workbook · Students: {student_rows:,} · Personas: {persona_rows:,} · "
                f"Phone keys: {len(phone_mapping):,} · Name keys: {len(name_mapping):,}"
            )

            with st.expander("Preview Sheet 1 - Students"):
                st.dataframe(students_df.head(20), use_container_width=True)

            with st.expander("Preview Sheet 2 - Personas"):
                st.dataframe(personas_df.head(20), use_container_width=True)

        st.markdown("---")

        admin_input = st.text_area(
            "Admin names / numbers",
            placeholder="Enter one admin name or number per line"
        )

        response_window_hours = st.slider(
            "Question response window",
            min_value=1,
            max_value=24,
            value=4
        )

        st.markdown("---")
        st.caption("Personas are highlighted in yellow wherever PersonType = Persona.")

    if not uploaded_files:
        st.info("Upload WhatsApp `.txt` exports to start.")
        return

    all_dfs = []

    with st.spinner("Parsing WhatsApp exports, matching students/personas, and detecting batch names..."):
        for file in uploaded_files:
            community_name = file.name.replace(".txt", "").replace("_", " ").strip()
            text = decode_file(file)
            parsed = parse_whatsapp_export(text, community_name)

            if parsed.empty:
                st.warning(f"No valid messages detected in: {file.name}")
            else:
                all_dfs.append(parsed)

    if not all_dfs:
        st.error("No valid WhatsApp messages found. Please check export format.")
        return

    df = pd.concat(all_dfs, ignore_index=True)
    df = df.sort_values("DateTime").reset_index(drop=True)

    df = enrich_with_people(df, phone_mapping, name_mapping)
    df = extract_features(df)
    df, rename_map = infer_community_names(df)

    admin_names = [x.strip() for x in admin_input.splitlines() if x.strip()]

    df = mark_answered_questions(
        df,
        admin_names=admin_names,
        response_window_hours=response_window_hours
    )

    min_date = df["Date"].min()
    max_date = df["Date"].max()

    st.markdown("---")

    top_left, top_right = st.columns([2, 1])

    with top_left:
        selected_range = st.date_input(
            "Select report date range",
            value=(max_date, max_date),
            min_value=min_date,
            max_value=max_date,
            format="DD/MM/YYYY"
        )

    with top_right:
        st.write("")
        st.write("")
        st.success(f"Loaded {len(uploaded_files)} communities · {len(df):,} parsed messages")

    with st.expander("Community auto-renaming based on dominant matched batch"):
        rename_df = pd.DataFrame(
            [{"Uploaded File Community": k, "Detected Community Name": v} for k, v in rename_map.items()]
        )
        st.dataframe(rename_df, use_container_width=True)

    if isinstance(selected_range, tuple):
        if len(selected_range) == 1:
            start_date = selected_range[0]
            end_date = selected_range[0]
        elif len(selected_range) >= 2:
            start_date = selected_range[0]
            end_date = selected_range[1]
        else:
            start_date = max_date
            end_date = max_date
    else:
        start_date = selected_range
        end_date = selected_range

    if start_date > end_date:
        st.error("Start date cannot be after end date.")
        return

    filtered_df = df[
        (df["Date"] >= start_date)
        & (df["Date"] <= end_date)
    ].copy()

    if filtered_df.empty:
        st.warning("No messages found for the selected date range.")
        return

    communities = sorted(filtered_df["Community"].dropna().unique().tolist())

    if not communities:
        st.warning("No communities found.")
        return

    tab_labels = communities + ["All Groups Comparison"]
    main_tabs = st.tabs(tab_labels)

    for i, community_name in enumerate(communities):
        with main_tabs[i]:
            community_df = filtered_df[filtered_df["Community"] == community_name].copy()

            render_community_page(
                community_name=community_name,
                community_df=community_df,
                admin_names=admin_names,
                start_date=start_date,
                end_date=end_date
            )

    with main_tabs[-1]:
        render_all_groups_comparison(
            filtered_df=filtered_df,
            admin_names=admin_names,
            start_date=start_date,
            end_date=end_date
        )


if __name__ == "__main__":
    main()
